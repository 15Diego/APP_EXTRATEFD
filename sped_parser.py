
import logging
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from dataclasses import dataclass
from enum import Enum
import time

import pandas as pd
import yaml

# Importa m√≥dulos customizados
from exceptions import (
    SpedError, SpedParseError, SpedValidationError, 
    SpedFileError, SpedEncodingError, SpedIntegrityError
)
from validators import (
    validate_cnpj, validate_date_format, validate_numeric_field,
    validate_registro, validate_cross_reference_totals,
    validate_chave_nfe, validate_cfop
)
from metrics import ProcessingMetrics

# Configura√ß√£o de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Carrega configura√ß√µes do arquivo YAML
CONFIG_FILE = Path(__file__).parent / 'config.yaml'
CONFIG = {}
try:
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            CONFIG = yaml.safe_load(f)
        logger.info(f"Configura√ß√µes carregadas de {CONFIG_FILE}")
    else:
        logger.warning(f"Arquivo de configura√ß√£o n√£o encontrado: {CONFIG_FILE}")
except Exception as e:
    logger.warning(f"Erro ao carregar configura√ß√µes: {e}. Usando valores padr√£o.")

# Fun√ß√£o auxiliar para obter configura√ß√µes com valor padr√£o
def get_config(path: str, default=None):
    """Obt√©m valor de configura√ß√£o usando nota√ß√£o de ponto (ex: 'processing.max_file_size_mb')"""
    keys = path.split('.')
    value = CONFIG
    for key in keys:
        if isinstance(value, dict):
            value = value.get(key)
        else:
            return default
    return value if value is not None else default


# =========================
# ENUMERA√á√ïES E CONSTANTES
# =========================

class IndicadorOperacao(Enum):
    """Indicador de tipo de opera√ß√£o."""
    ENTRADA = '0'
    SAIDA = '1'


class IndicadorEmitente(Enum):
    """Indicador de emitente do documento."""
    EMISSAO_PROPRIA = '0'
    TERCEIROS = '1'


class IndicadorFrete(Enum):
    """Indicador de tipo de frete."""
    EMITENTE = '0'
    DESTINATARIO_REMETENTE = '1'
    TERCEIROS = '2'
    SEM_COBRANCA = '9'


# Mapeamentos para convers√£o de c√≥digos
IND_OPER_MAP = {
    '0': 'Entrada',
    '1': 'Sa√≠da'
}

IND_EMIT_MAP = {
    '0': 'Emiss√£o pr√≥pria',
    '1': 'Terceiros'
}

IND_FRT_MAP = {
    '0': 'Emitente',
    '1': 'Destinat√°rio/remetente',
    '2': 'Terceiros',
    '9': 'Sem cobran√ßa de frete'
}

# Formato de moeda brasileira para Excel
BRL_ACCOUNTING = get_config('export.currency_format', 
    '_-* "R$" * #,##0.00_-;_-* "R$" * -#,##0.00_-;_-* "R$" * "-"??_-;_-@_-')

# Tamanho m√°ximo de arquivo (em MB, convertido para bytes)
MAX_FILE_SIZE = get_config('processing.max_file_size_mb', 100) * 1024 * 1024

# Toler√¢ncia para valida√ß√£o de totais
VALIDATION_TOLERANCE = get_config('processing.validation_tolerance', 0.01)


# =========================
# LAYOUTS DOS REGISTROS
# =========================

@dataclass
class RegistroLayout:
    """Define o layout de um registro SPED."""
    codigo: str
    campos: List[str]
    campos_numericos: List[str]


# Layouts dos principais registros SPED
LAYOUTS = {
    # --- Bloco C: Documentos Fiscais I ---
    'C010': ['REG', 'CNPJ', 'IND_ESCRI'],
    'C100': [
        'REG', 'IND_OPER', 'IND_EMIT', 'COD_PART', 'COD_MOD', 'COD_SIT', 'SER', 'NUM_DOC',
        'CHV_NFE', 'DT_DOC', 'DT_E_S', 'VL_DOC', 'IND_PGTO', 'VL_DESC', 'VL_ABAT_NT',
        'VL_MERC', 'IND_FRT', 'VL_FRT', 'VL_SEG', 'VL_OUT_DA', 'VL_BC_ICMS', 'VL_ICMS',
        'VL_BC_ICMS_ST', 'VL_ICMS_ST', 'VL_IPI', 'VL_PIS', 'VL_COFINS', 'VL_PIS_ST', 'VL_COFINS_ST'
    ],
    'C170': [
        'REG', 'NUM_ITEM', 'COD_ITEM', 'DESCR_COMPL', 'QTD', 'UNID', 'VL_ITEM', 'VL_DESC',
        'IND_MOV', 'CST_ICMS', 'CFOP', 'COD_NAT', 'VL_BC_ICMS', 'ALIQ_ICMS', 'VL_ICMS',
        'VL_BC_ICMS_ST', 'ALIQ_ST', 'VL_ICMS_ST', 'IND_APUR', 'CST_IPI', 'COD_ENQ',
        'VL_BC_IPI', 'ALIQ_IPI', 'VL_IPI', 'CST_PIS', 'VL_BC_PIS', 'ALIQ_PIS', 'QUANT_BC_PIS',
        'ALIQ_PIS_QUANT', 'VL_PIS', 'CST_COFINS', 'VL_BC_COFINS', 'ALIQ_COFINS', 'QUANT_BC_COFINS',
        'ALIQ_COFINS_QUANT', 'VL_COFINS', 'COD_CTA'
    ],
    'C190': [
        'REG', 'CST_ICMS', 'CFOP', 'ALIQ_ICMS', 'VL_OPR', 'VL_BC_ICMS', 'VL_ICMS',
        'VL_BC_ICMS_ST', 'VL_ICMS_ST', 'VL_RED_BC', 'VL_IPI', 'COD_OBS'
    ],
    'C195': ['REG', 'COD_OBS', 'TXT_COMPL'],
    'C197': [
        'REG', 'COD_AJ', 'DESCR_COMPL_AJ', 'COD_ITEM', 'VL_BC_ICMS',
        'ALIQ_ICMS', 'VL_ICMS', 'VL_OUTROS'
    ],

    # --- Bloco C500: Notas Fiscais de Energia El√©trica/Servi√ßos (ICMS) ---
    # Registro principal C500 e seus complementos C501 (PIS) e C505 (COFINS)
    'C500': [
        'REG', 'COD_PART', 'COD_MOD', 'COD_SIT', 'SER', 'SUB', 'NUM_DOC',
        'DT_DOC', 'DT_ENT', 'VL_DOC', 'VL_ICMS', 'COD_INF', 'VL_PIS',
        'VL_COFINS', 'CHV_DOCe'
    ],
    # Detalhamento de PIS para C500
    'C501': [
        'REG', 'CST_PIS', 'VL_ITEM', 'NAT_BC_CRED', 'VL_BC_PIS',
        'ALIQ_PIS', 'VL_PIS', 'COD_CTA'
    ],
    # Detalhamento de COFINS para C500
    'C505': [
        'REG', 'CST_COFINS', 'VL_ITEM', 'NAT_BC_CRED', 'VL_BC_COFINS',
        'ALIQ_COFINS', 'VL_COFINS', 'COD_CTA'
    ],

    # --- Bloco D: Documentos Fiscais II ---
    'D010': ['REG', 'CNPJ'],
    'D100': [
        'REG', 'IND_OPER', 'IND_EMIT', 'COD_PART', 'COD_MOD', 'COD_SIT', 'SER', 'SUB', 'NUM_DOC',
        'CHV_CTE', 'DT_DOC', 'DT_A_P', 'TP_CT_E', 'CHV_CTE_REF', 'VL_DOC', 'VL_DESC', 'IND_FRT',
        'VL_SERV', 'VL_BC_ICMS', 'VL_ICMS', 'VL_NT', 'COD_INF', 'COD_CTA', 'COD_MUN_ORIG', 'COD_MUN_DEST'
    ],
    'D170': [
        'REG', 'COD_ITEM', 'DESCR_COMPL', 'QTD', 'UNID', 'VL_ITEM', 'VL_DESC', 'IND_MOV',
        'CST_ICMS', 'CFOP', 'COD_NAT', 'VL_BC_ICMS', 'ALIQ_ICMS', 'VL_ICMS', 'VL_BC_ICMS_ST',
        'ALIQ_ST', 'VL_ICMS_ST', 'IND_APUR', 'COD_CTA'
    ],
    'D190': [
        'REG', 'CST_ICMS', 'CFOP', 'ALIQ_ICMS', 'VL_OPR', 'VL_BC_ICMS',
        'VL_ICMS', 'VL_RED_BC', 'COD_OBS'
    ],
    'D101': [
        'REG', 'IND_NAT_FRT', 'VL_ITEM', 'CST_PIS', 'NAT_BC_CRED',
        'VL_BC_PIS', 'ALIQ_PIS', 'VL_PIS', 'COD_CTA'
    ],
    'D105': [
        'REG', 'IND_NAT_FRT', 'VL_ITEM', 'CST_COFINS', 'NAT_BC_CRED',
        'VL_BC_COFINS', 'ALIQ_COFINS', 'VL_COFINS', 'COD_CTA'
    ],

    # --- Bloco D500: Notas de Servi√ßos de Comunica√ß√£o/Telecom (ICMS) ---
    # Registro principal D500 com seus complementos D501 (PIS) e D505 (COFINS)
    'D500': [
        'REG', 'IND_OPER', 'IND_EMIT', 'COD_PART', 'COD_MOD', 'COD_SIT', 'SER',
        'SUB', 'NUM_DOC', 'DT_DOC', 'DT_A_P', 'VL_DOC', 'VL_DESC', 'VL_SERV',
        'VL_SERV_NT', 'VL_TERC', 'VL_DA', 'VL_BC_ICMS', 'VL_ICMS', 'COD_INF',
        'VL_PIS', 'VL_COFINS', 'CHV_DOCe'
    ],
    # Detalhamento de PIS para D500
    'D501': [
        'REG', 'CST_PIS', 'VL_ITEM', 'NAT_BC_CRED', 'VL_BC_PIS',
        'ALIQ_PIS', 'VL_PIS', 'COD_CTA'
    ],
    # Detalhamento de COFINS para D500
    'D505': [
        'REG', 'CST_COFINS', 'VL_ITEM', 'NAT_BC_CRED', 'VL_BC_COFINS',
        'ALIQ_COFINS', 'VL_COFINS', 'COD_CTA'
    ],

    # --- Bloco D700: NFCom ‚Äì Nota Fiscal Fatura de Servi√ßos de Comunica√ß√£o Eletr√¥nica ---
    'D700': [
        'REG', 'IND_OPER', 'IND_EMIT', 'COD_PART', 'COD_MOD', 'COD_SIT', 'SER',
        'NUM_DOC', 'DT_DOC', 'DT_E_S', 'VL_DOC', 'VL_DESC', 'VL_SERV',
        'VL_SERV_NT', 'VL_TERC', 'VL_DA', 'VL_BC_ICMS', 'VL_ICMS', 'COD_INF',
        'VL_PIS', 'VL_COFINS', 'CHV_DOCe', 'FIN_DOCe', 'TIP_FAT', 'COD_MOD_DOC_REF',
        'CHV_DOCe_REF', 'HASH_DOC_REF', 'SER_DOC_REF', 'NUM_DOC_REF',
        'MES_DOC_REF', 'COD_MUN_DEST', 'DED'
    ],

    # --- Bloco A: Documentos Fiscais III ---
    'A001': ['REG', 'IND_MOV'],
    'A010': ['REG', 'CNPJ'],
    'A100': [
        'REG', 'IND_OPER', 'IND_EMIT', 'COD_PART', 'COD_SIT', 'SER', 'SUB', 'NUM_DOC',
        'CHV_NFSE', 'DT_DOC', 'DT_EXE_SERV', 'VL_DOC', 'IND_PGTO', 'VL_DESC', 'VL_BC_PIS',
        'VL_PIS', 'VL_BC_COFINS', 'VL_COFINS', 'VL_PIS_RET', 'VL_COFINS_RET', 'VL_ISS'
    ],

    # --- Bloco F: Demais Documentos ---
    'F001': ['REG', 'IND_MOV'],
    'F010': ['REG', 'CNPJ'],
    'F100': [
        'REG', 'IND_OPER', 'COD_PART', 'COD_ITEM', 'DT_OPER', 'VL_OPER', 'CST_PIS',
        'VL_BC_PIS', 'ALIQ_PIS', 'VL_PIS', 'CST_COFINS', 'VL_BC_COFINS', 'ALIQ_COFINS',
        'VL_COFINS', 'NAT_BC_CRED', 'IND_ORIG_CRED', 'COD_CTA', 'COD_CCUS', 'DESC_COMPL'
    ],
    'F111': ['REG', 'NUM_PROC', 'IND_PROC'],

    # --- Bloco M: Apura√ß√£o da Contribui√ß√£o ---
    'M001': ['REG', 'IND_MOV'],
    'M100': [
        'REG', 'COD_CRED', 'IND_CRED_ORI', 'VL_BC_PIS', 'ALIQ_PIS', 'QUANT_BC_PIS',
        'ALIQ_PIS_QUANT', 'VL_CRED', 'VL_AJUS_ACRES', 'VL_AJUS_REDUC', 'VL_CRED_DIF',
        'VL_CRED_DISP', 'IND_DESC_CRED', 'VL_CRED_DESC', 'SLD_CRED'
    ],
    'M105': [
        'REG', 'NAT_BC_CRED', 'CST_PIS', 'VL_BC_PIS_TOT', 'VL_BC_PIS_CUM', 'VL_BC_PIS_NC',
        'VL_BC_PIS', 'QUANT_BC_PIS_TOT', 'QUANT_BC_PIS', 'DESC_CRED'
    ],
    'M110': ['REG', 'IND_AJ', 'VL_AJ', 'COD_AJ', 'NUM_DOC', 'DESCR_AJ', 'DT_REF'],
    'M115': [
        'REG', 'DET_VALOR_AJ', 'CST_PIS', 'DET_BC_CRED', 'DET_ALIQ',
        'DT_OPER_AJ', 'DESC_AJ', 'COD_CTA', 'INFO_COMPL'
    ],

    # --- Bloco E: Apura√ß√£o do ICMS/IPI ---
    # E001: Abertura do Bloco E
    'E001': ['REG', 'IND_DAD'],
    # E100: Per√≠odo da apura√ß√£o
    'E100': ['REG', 'DT_INI', 'DT_FIN'],
    # E110: Apura√ß√£o do ICMS ‚Äì Opera√ß√µes pr√≥prias
    'E110': [
        'REG', 'VL_TOT_DEBITOS', 'VL_AJ_DEBITOS', 'VL_TOT_AJ_DEBITOS',
        'VL_ESTORNOS_CRED', 'VL_TOT_CREDITOS', 'VL_AJ_CREDITOS', 'VL_TOT_AJ_CREDITOS',
        'VL_ESTORNOS_DEB', 'VL_SLD_CREDOR_ANT', 'VL_SLD_APURADO', 'VL_TOT_DED',
        'VL_ICMS_RECOLHER', 'VL_SLD_CREDOR_TRANSPORTAR', 'DEB_ESP'
    ],
    # E111: Ajustes da apura√ß√£o do ICMS
    'E111': ['REG', 'COD_AJ_APUR', 'DESCR_COMPL_AJ', 'VL_AJ_APUR'],
    # E112: Informa√ß√µes adicionais dos ajustes da apura√ß√£o do ICMS
    'E112': ['REG', 'NUM_DA', 'NUM_PROC', 'IND_PROC', 'PROC', 'COD_OBS'],
    # E113: Informa√ß√µes adicionais dos ajustes da apura√ß√£o do ICMS ‚Äì Identifica√ß√£o dos documentos fiscais
    'E113': ['REG', 'COD_PART', 'COD_MOD', 'SER', 'SUB', 'NUM_DOC', 'DT_DOC', 'COD_ITEM', 'VL_AJ_ITEM', 'CHV_DOCe'],
    # E115: Informa√ß√µes adicionais da apura√ß√£o ‚Äì valores declarat√≥rios
    'E115': ['REG', 'COD_INF_ADIC', 'VL_INF_ADIC', 'DESCR_COMPL_AJ'],
    # E116: Obriga√ß√µes do ICMS a recolher ‚Äì opera√ß√µes pr√≥prias
    'E116': ['REG', 'COD_OR', 'VL_OR', 'DT_VCTO', 'COD_REC', 'NUM_PROC', 'IND_PROC', 'PROC', 'TXT_COMPL', 'MES_REF'],
}

# Campos num√©ricos por registro
NUMERIC_COLUMNS = {
    'C100': [
        'VL_DOC', 'VL_DESC', 'VL_ABAT_NT', 'VL_MERC', 'VL_FRT', 'VL_SEG', 'VL_OUT_DA',
        'VL_BC_ICMS', 'VL_ICMS', 'VL_BC_ICMS_ST', 'VL_ICMS_ST', 'VL_IPI', 'VL_PIS',
        'VL_COFINS', 'VL_PIS_ST', 'VL_COFINS_ST'
    ],
    'C170': [
        'QTD', 'VL_ITEM', 'VL_DESC', 'VL_BC_ICMS', 'ALIQ_ICMS', 'VL_ICMS', 'VL_BC_ICMS_ST',
        'ALIQ_ST', 'VL_ICMS_ST', 'VL_BC_IPI', 'ALIQ_IPI', 'VL_IPI', 'VL_BC_PIS', 'ALIQ_PIS',
        'QUANT_BC_PIS', 'ALIQ_PIS_QUANT', 'VL_PIS', 'VL_BC_COFINS', 'ALIQ_COFINS',
        'QUANT_BC_COFINS', 'ALIQ_COFINS_QUANT', 'VL_COFINS'
    ],
    'C190': [
        'ALIQ_ICMS', 'VL_OPR', 'VL_BC_ICMS', 'VL_ICMS', 'VL_BC_ICMS_ST',
        'VL_ICMS_ST', 'VL_RED_BC', 'VL_IPI'
    ],
    'C197': ['VL_BC_ICMS', 'ALIQ_ICMS', 'VL_ICMS', 'VL_OUTROS'],
    'D100': ['VL_DOC', 'VL_DESC', 'VL_SERV', 'VL_BC_ICMS', 'VL_ICMS', 'VL_NT'],
    'D170': [
        'QTD', 'VL_ITEM', 'VL_DESC', 'VL_BC_ICMS', 'ALIQ_ICMS', 'VL_ICMS',
        'VL_BC_ICMS_ST', 'ALIQ_ST', 'VL_ICMS_ST'
    ],
    'D190': ['ALIQ_ICMS', 'VL_OPR', 'VL_BC_ICMS', 'VL_ICMS', 'VL_RED_BC'],
    'D101': ['VL_ITEM', 'VL_BC_PIS', 'ALIQ_PIS', 'VL_PIS'],
    'D105': ['VL_ITEM', 'VL_BC_COFINS', 'ALIQ_COFINS', 'VL_COFINS'],
    'A100': [
        'VL_DOC', 'VL_DESC', 'VL_BC_PIS', 'VL_PIS', 'VL_BC_COFINS',
        'VL_COFINS', 'VL_PIS_RET', 'VL_COFINS_RET', 'VL_ISS'
    ],
    'F100': [
        'VL_OPER', 'VL_BC_PIS', 'ALIQ_PIS', 'VL_PIS', 'VL_BC_COFINS',
        'ALIQ_COFINS', 'VL_COFINS'
    ],
    'M100': [
        'VL_BC_PIS', 'ALIQ_PIS', 'QUANT_BC_PIS', 'ALIQ_PIS_QUANT', 'VL_CRED',
        'VL_AJUS_ACRES', 'VL_AJUS_REDUC', 'VL_CRED_DIF', 'VL_CRED_DISP',
        'VL_CRED_DESC', 'SLD_CRED'
    ],
    'M105': [
        'VL_BC_PIS_TOT', 'VL_BC_PIS_CUM', 'VL_BC_PIS_NC', 'VL_BC_PIS',
        'QUANT_BC_PIS_TOT', 'QUANT_BC_PIS'
    ],
    'M110': ['VL_AJ'],
    'M115': ['DET_VALOR_AJ', 'DET_BC_CRED', 'DET_ALIQ'],

    # Bloco E
    'E110': [
        'VL_TOT_DEBITOS', 'VL_AJ_DEBITOS', 'VL_TOT_AJ_DEBITOS',
        'VL_ESTORNOS_CRED', 'VL_TOT_CREDITOS', 'VL_AJ_CREDITOS', 'VL_TOT_AJ_CREDITOS',
        'VL_ESTORNOS_DEB', 'VL_SLD_CREDOR_ANT', 'VL_SLD_APURADO', 'VL_TOT_DED',
        'VL_ICMS_RECOLHER', 'VL_SLD_CREDOR_TRANSPORTAR', 'DEB_ESP'
    ],
    'E111': ['VL_AJ_APUR'],
    'E113': ['VL_AJ_ITEM'],
    'E115': ['VL_INF_ADIC'],
    'E116': ['VL_OR'],

    # Valores monet√°rios para C500 e seus complementos
    'C500': ['VL_DOC', 'VL_ICMS', 'VL_PIS', 'VL_COFINS'],
    'C501': ['VL_ITEM', 'VL_BC_PIS', 'ALIQ_PIS', 'VL_PIS'],
    'C505': ['VL_ITEM', 'VL_BC_COFINS', 'ALIQ_COFINS', 'VL_COFINS'],

    # Valores monet√°rios para D500 e seus complementos
    'D500': [
        'VL_DOC', 'VL_DESC', 'VL_SERV', 'VL_SERV_NT', 'VL_TERC', 'VL_DA',
        'VL_BC_ICMS', 'VL_ICMS', 'VL_PIS', 'VL_COFINS'
    ],
    'D501': ['VL_ITEM', 'VL_BC_PIS', 'ALIQ_PIS', 'VL_PIS'],
    'D505': ['VL_ITEM', 'VL_BC_COFINS', 'ALIQ_COFINS', 'VL_COFINS'],

    # Valores monet√°rios para D700
    'D700': [
        'VL_DOC', 'VL_DESC', 'VL_SERV', 'VL_SERV_NT', 'VL_TERC', 'VL_DA',
        'VL_BC_ICMS', 'VL_ICMS', 'VL_PIS', 'VL_COFINS', 'DED'
    ],
}

# Defini√ß√£o de grupos de consolida√ß√£o (pai, filhos, √≠ndice_pai, √≠ndice_header, registro_header)
GROUPS: Dict[str, Tuple[str, List[str], str, str, str]] = {
    'C': ('C100', ['C170', 'C190', 'C195', 'C197'], 'C100_INDEX', 'C010_INDEX', 'C010'),
    'D': ('D100', ['D170', 'D190', 'D101', 'D105'], 'D100_INDEX', 'D010_INDEX', 'D010'),
    'A': ('A100', [], 'A100_INDEX', 'A010_INDEX', 'A010'),
    'F': ('F100', ['F111'], 'F100_INDEX', 'F010_INDEX', 'F010'),

    # Bloco E: apura√ß√£o do ICMS/IPI (Bloco Fiscal)
    'E': ('E110', ['E111', 'E112', 'E113', 'E115', 'E116'], 'E110_INDEX', 'E100_INDEX', 'E100'),

    # Bloco C500: notas fiscais de energia el√©trica e servi√ßos (entradas) com cr√©ditos
    'C500': ('C500', ['C501', 'C505'], 'C500_INDEX', 'C010_INDEX', 'C010'),

    # Bloco D500: notas fiscais de servi√ßos de comunica√ß√£o/telecom (aquisi√ß√£o) com direito a cr√©dito
    'D500': ('D500', ['D501', 'D505'], 'D500_INDEX', 'D010_INDEX', 'D010'),

    # Bloco D700: NFCom ‚Äì nota fiscal fatura eletr√¥nica de servi√ßos de comunica√ß√£o
    # N√£o possui filhos (at√© o layout atual)
    'D700': ('D700', [], 'D700_INDEX', 'D010_INDEX', 'D010'),
}


# =========================
# FUN√á√ïES UTILIT√ÅRIAS
# =========================

def detect_encoding(file_path: Path, sample_bytes: int = None) -> str:
    """
    Detecta o encoding de um arquivo automaticamente.

    Args:
        file_path: Caminho do arquivo
        sample_bytes: N√∫mero de bytes a ler para detec√ß√£o (None = usar config)

    Returns:
        Nome do encoding detectado
        
    Raises:
        SpedEncodingError: Se n√£o conseguir detectar encoding v√°lido
    """
    if sample_bytes is None:
        sample_bytes = get_config('processing.encoding_sample_bytes', 256_000)
    
    try:
        raw_data = file_path.read_bytes()[:sample_bytes]
        
        # Tenta usar charset_normalizer se dispon√≠vel
        try:
            from charset_normalizer import from_bytes
            result = from_bytes(raw_data).best()
            if result and result.encoding:
                logger.info(f"Encoding detectado via charset_normalizer: {result.encoding}")
                return result.encoding
        except ImportError:
            logger.debug("charset_normalizer n√£o dispon√≠vel, usando fallback")
        
        # Tenta encodings de fallback
        fallback_encodings = get_config('processing.fallback_encodings', 
                                       ['latin-1', 'utf-8', 'cp1252', 'iso-8859-1'])
        
        for encoding in fallback_encodings:
            try:
                raw_data.decode(encoding)
                logger.info(f"Encoding detectado via fallback: {encoding}")
                return encoding
            except (UnicodeDecodeError, LookupError):
                continue
        
        # Se nenhum encoding funcionou, usa padr√£o
        default_encoding = get_config('processing.default_encoding', 'latin-1')
        logger.warning(f"Nenhum encoding detectado, usando padr√£o: {default_encoding}")
        return default_encoding
    
    except Exception as e:
        logger.error(f"Erro ao detectar encoding: {e}")
        raise SpedEncodingError(f"Falha ao detectar encoding", str(file_path)) from e


def parse_sped_line(line: str) -> List[str]:
    """
    Faz o parse de uma linha SPED, removendo pipes inicial e final.

    Args:
        line: Linha do arquivo SPED

    Returns:
        Lista de campos da linha
    """
    parts = line.rstrip('\r\n').split('|')
    
    # Remove pipe inicial se existir
    if parts and parts[0] == '':
        parts = parts[1:]
    
    # Remove pipe final se existir
    if parts and parts[-1] == '':
        parts = parts[:-1]
    
    return parts


def convert_numeric_columns(df: pd.DataFrame, columns: List[str]) -> None:
    """
    Converte colunas num√©ricas do formato brasileiro para formato Python.
    
    Formato brasileiro: 1.234,56
    Formato Python: 1234.56
    
    Otimizado com opera√ß√µes vetorizadas do pandas.

    Args:
        df: DataFrame a ser modificado (in-place)
        columns: Lista de colunas a converter
    """
    for col in columns:
        if col not in df.columns:
            continue
        
        try:
            # Converte para string apenas se necess√°rio
            if not pd.api.types.is_string_dtype(df[col]):
                df[col] = df[col].astype(str)
            
            # Opera√ß√µes vetorizadas: remove pontos e substitui v√≠rgula por ponto
            # Usa regex=False para melhor performance
            df[col] = (
                df[col]
                .str.replace('.', '', regex=False)
                .str.replace(',', '.', regex=False)
                .str.strip()
            )
            
            # Converte para num√©rico
            # errors='coerce' transforma valores inv√°lidos em NaN
            df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # Log de valores inv√°lidos se houver muitos
            null_count = df[col].isna().sum()
            if null_count > 0:
                logger.debug(f"Coluna {col}: {null_count} valores inv√°lidos convertidos para NaN")
                
        except Exception as e:
            logger.warning(f"Erro ao converter coluna {col}: {e}")


def concat_unique_values(series: pd.Series) -> str:
    """
    Concatena valores √∫nicos de uma s√©rie, separados por ponto e v√≠rgula.

    Args:
        series: S√©rie pandas com valores a concatenar

    Returns:
        String com valores √∫nicos concatenados
    """
    values = [str(v) for v in series if pd.notna(v) and str(v).strip()]
    unique_values = sorted(set(values))
    return '; '.join(unique_values) if unique_values else ''


def validate_file_path(file_path: Path) -> None:
    """
    Valida se o caminho do arquivo √© v√°lido e seguro.

    Args:
        file_path: Caminho do arquivo

    Raises:
        SpedFileError: Se o arquivo for inv√°lido, n√£o existir, ou for muito grande
    """
    try:
        # Sanitiza o caminho para prevenir path traversal
        file_path = file_path.resolve().absolute()
        
        # Verifica se o arquivo existe
        if not file_path.exists():
            raise SpedFileError("Arquivo n√£o encontrado", str(file_path))
        
        # Verifica se √© um arquivo (n√£o diret√≥rio)
        if not file_path.is_file():
            raise SpedFileError("Caminho n√£o √© um arquivo", str(file_path))
        
        # Verifica tamanho do arquivo
        file_size = file_path.stat().st_size
        if file_size == 0:
            raise SpedFileError("Arquivo vazio", str(file_path))
        
        if file_size > MAX_FILE_SIZE:
            max_mb = MAX_FILE_SIZE / (1024*1024)
            actual_mb = file_size / (1024*1024)
            raise SpedFileError(
                f"Arquivo muito grande: {actual_mb:.2f} MB (m√°ximo: {max_mb:.0f} MB)",
                str(file_path)
            )
        
        # Verifica extens√£o
        if file_path.suffix.lower() not in ['.txt', '.sped', '']:
            logger.warning(f"Extens√£o incomum: {file_path.suffix}")
        
        logger.info(f"Arquivo validado: {file_path.name} ({file_size / 1024:.1f} KB)")
    
    except SpedFileError:
        raise
    except Exception as e:
        logger.error(f"Erro ao validar arquivo: {e}")
        raise SpedFileError(f"Erro na valida√ß√£o do arquivo: {e}", str(file_path)) from e


# =========================
# PARSER DE ARQUIVOS SPED
# =========================

class SpedParser:
    """Parser para arquivos SPED."""
    
    def __init__(self, file_path: Path, layouts: Dict[str, List[str]] = None, 
                 numeric_columns: Dict[str, List[str]] = None,
                 groups: Dict[str, Tuple[str, List[str], str, str, str]] = None):
        """
        Inicializa o parser.

        Args:
            file_path: Caminho do arquivo SPED
            layouts: Dicion√°rio opcional de layouts por registro.
            numeric_columns: Dicion√°rio opcional de colunas num√©ricas por registro.
            groups: Dicion√°rio opcional de grupos de consolida√ß√£o.
            
        Raises:
            SpedFileError: Se o arquivo for inv√°lido
            SpedEncodingError: Se n√£o conseguir detectar encoding
        """
        self.file_path = file_path
        self.encoding = detect_encoding(file_path)
        
        # Usa configura√ß√µes externas ou internas
        self.layouts = layouts if layouts is not None else LAYOUTS
        self.numeric_columns = numeric_columns if numeric_columns is not None else NUMERIC_COLUMNS
        self.groups = groups if groups is not None else GROUPS
        
        self.rows: Dict[str, List[List[str]]] = {code: [] for code in self.layouts}
        
        # M√©tricas de processamento
        self.metrics = ProcessingMetrics()
        self.metrics.arquivo_processado = str(file_path.name)
        
        # √çndices de controle por bloco
        self.indices = {
            'c010': -1, 'c100': -1, 'c500': -1,
            'd010': -1, 'd100': -1, 'd500': -1, 'd700': -1,
            'a010': -1, 'a100': -1,
            'f010': -1, 'f100': -1,
            'e100': -1, 'e110': -1,
            'h001': -1, 'h005': -1,
            'k001': -1, 'k100': -1,
            'g001': -1, 'g110': -1,
            'm001': -1, 'm100': -1, 'm500': -1,
            '0000': -1,
        }
        
        # Mapeamento para processamento de √≠ndices gen√©ricos
        # RecordCode -> Lista de a√ß√µes ({type: increment/read, key: index_key, col: col_name})
        self.record_actions = {}
        
        # Inicializa a√ß√µes a partir dos grupos
        for group_info in self.groups.values():
            parent, children, parent_idx_name, header_idx_name, header = group_info
            
            # --- CONFIGURA√á√ÉO DO PAI ---
            if parent not in self.record_actions:
                self.record_actions[parent] = []
            
            parent_key = parent.lower()
            
            # Pai incrementa seu pr√≥prio √≠ndice
            if not any(a['type'] == 'increment' and a['key'] == parent_key for a in self.record_actions[parent]):
                self.record_actions[parent].append({
                    'type': 'increment',
                    'key': parent_key,
                    'col': parent_idx_name
                })
            
            # Pai l√™ √≠ndice do header (se existir e for diferente)
            if header and header != parent:
                header_key = header.lower()
                if not any(a['type'] == 'read' and a['key'] == header_key for a in self.record_actions[parent]):
                    self.record_actions[parent].append({
                        'type': 'read',
                        'key': header_key,
                        'col': header_idx_name
                    })
            
            # --- CONFIGURA√á√ÉO DO HEADER ---
            if header:
                if header not in self.record_actions:
                    self.record_actions[header] = []
                
                header_key = header.lower()
                # Header incrementa seu pr√≥prio √≠ndice
                if not any(a['type'] == 'increment' and a['key'] == header_key for a in self.record_actions[header]):
                    self.record_actions[header].append({
                        'type': 'increment',
                        'key': header_key,
                        'col': header_idx_name
                    })

            # --- CONFIGURA√á√ÉO DOS FILHOS ---
            for child in children:
                if child not in self.record_actions:
                    self.record_actions[child] = []
                
                # Filho l√™ √≠ndice do pai
                if not any(a['type'] == 'read' and a['key'] == parent_key for a in self.record_actions[child]):
                    self.record_actions[child].append({
                        'type': 'read',
                        'key': parent_key,
                        'col': parent_idx_name
                    })
        
        # Configura√ß√µes de valida√ß√£o
        self.validate_data = get_config('validation.validate_required_fields', True)
        self.strict_mode = get_config('validation.strict_mode', False)
    
    def parse(self) -> Dict[str, pd.DataFrame]:
        """
        Faz o parse do arquivo SPED completo.

        Returns:
            Dicion√°rio com DataFrames para cada registro
            
        Raises:
            SpedParseError: Se houver erro cr√≠tico no parsing
        """
        logger.info(f"Iniciando parse do arquivo: {self.file_path}")
        start_time = time.time()
        
        try:
            with self.file_path.open('r', encoding=self.encoding, errors='replace') as file:
                # Conta total de linhas primeiro (para m√©tricas)
                file.seek(0)
                self.metrics.total_lines = sum(1 for _ in file)
                file.seek(0)
                
                for line_number, raw_line in enumerate(file, 1):
                    # Ignora linhas vazias ou que n√£o come√ßam com pipe
                    if not raw_line.strip() or not raw_line.startswith('|'):
                        self.metrics.skipped_lines += 1
                        continue
                    
                    try:
                        self._process_line(raw_line, line_number)
                    except SpedParseError as e:
                        self.metrics.increment_erro("Parse Error")
                        if self.strict_mode:
                            raise
                        logger.warning(str(e))
                    except Exception as e:
                        self.metrics.increment_erro("Erro Desconhecido")
                        logger.warning(f"Erro na linha {line_number}: {e}")
                        if self.strict_mode:
                            raise SpedParseError(
                                f"Erro inesperado ao processar linha",
                                line_number=line_number,
                                line_content=raw_line
                            ) from e
            
            # Converte listas em DataFrames
            dataframes = self._create_dataframes()
            
            # Finaliza m√©tricas
            self.metrics.finalizar()
            self.metrics.log_summary()
            
            logger.info(f"Parse conclu√≠do em {time.time() - start_time:.2f}s. "
                       f"{len(dataframes)} tipos de registro processados.")
            
            return dataframes
        
        except Exception as e:
            logger.error(f"Erro ao fazer parse do arquivo: {e}")
            raise
    
    def _process_line(self, raw_line: str, line_number: int = None) -> None:
        """
        Processa uma linha do arquivo SPED.

        Args:
            raw_line: Linha bruta do arquivo
            line_number: N√∫mero da linha (para mensagens de erro)
            
        Raises:
            SpedParseError: Se a linha estiver malformada
        """
        # Extrai c√≥digo do registro
        if len(raw_line) < 5:
            raise SpedParseError(
                "Linha muito curta para conter registro v√°lido",
                line_number=line_number,
                line_content=raw_line
            )
        
        registro = raw_line[1:5]
        
        # Incrementa m√©trica
        self.metrics.increment_registro(registro)
        
        # Processa registro usando layouts din√¢micos
        if registro in self.layouts:
            self._process_generic(registro, raw_line)
        # Registros desconhecidos s√£o ignorados (mas j√° contados na m√©trica)
    
    def _process_c010(self, raw_line: str) -> None:
        """Processa registro C010."""
        self.indices['c010'] += 1
        parts = self._pad_line(raw_line, 'C010')
        parts.append(self.indices['c010'])
        self.rows['C010'].append(parts)
    
    def _process_c100(self, raw_line: str) -> None:
        """Processa registro C100."""
        self.indices['c100'] += 1
        parts = self._pad_line(raw_line, 'C100')
        parts.extend([self.indices['c100'], self.indices['c010']])
        self.rows['C100'].append(parts)
    
    def _process_c_child(self, registro: str, raw_line: str) -> None:
        """Processa registros filhos do C100 (C170, C190, C195, C197)."""
        parts = self._pad_line(raw_line, registro)
        parts.append(self.indices['c100'])
        self.rows[registro].append(parts)
    
    def _process_d010(self, raw_line: str) -> None:
        """Processa registro D010."""
        self.indices['d010'] += 1
        parts = self._pad_line(raw_line, 'D010')
        parts.append(self.indices['d010'])
        self.rows['D010'].append(parts)
    
    def _process_d100(self, raw_line: str) -> None:
        """Processa registro D100."""
        self.indices['d100'] += 1
        parts = self._pad_line(raw_line, 'D100')
        parts.extend([self.indices['d100'], self.indices['d010']])
        self.rows['D100'].append(parts)
    
    def _process_d_child(self, registro: str, raw_line: str) -> None:
        """Processa registros filhos do D100 (D170, D190, D101, D105)."""
        parts = self._pad_line(raw_line, registro)
        parts.append(self.indices['d100'])
        self.rows[registro].append(parts)
    
    def _process_a010(self, raw_line: str) -> None:
        """Processa registro A010."""
        self.indices['a010'] += 1
        parts = self._pad_line(raw_line, 'A010')
        parts.append(self.indices['a010'])
        self.rows['A010'].append(parts)
    
    def _process_a100(self, raw_line: str) -> None:
        """Processa registro A100."""
        self.indices['a100'] += 1
        parts = self._pad_line(raw_line, 'A100')
        parts.extend([self.indices['a100'], self.indices['a010']])
        self.rows['A100'].append(parts)
    
    def _process_f010(self, raw_line: str) -> None:
        """Processa registro F010."""
        self.indices['f010'] += 1
        parts = self._pad_line(raw_line, 'F010')
        parts.append(self.indices['f010'])
        self.rows['F010'].append(parts)
    
    def _process_f100(self, raw_line: str) -> None:
        """Processa registro F100."""
        self.indices['f100'] += 1
        parts = self._pad_line(raw_line, 'F100')
        parts.extend([self.indices['f100'], self.indices['f010']])
        self.rows['F100'].append(parts)
    
    def _process_f_child(self, registro: str, raw_line: str) -> None:
        """Processa registros filhos do F100."""
        parts = self._pad_line(raw_line, registro)
        parts.append(self.indices['f100'])
        self.rows[registro].append(parts)
    
    def _process_m(self, registro: str, raw_line: str) -> None:
        """Processa registros do Bloco M."""
        parts = self._pad_line(raw_line, registro)
        self.rows[registro].append(parts)

    # ======== Bloco E ========
    def _process_e100(self, raw_line: str) -> None:
        """
        Processa registro E100 (per√≠odo de apura√ß√£o).

        Atribui um √≠ndice sequencial e armazena o registro.
        """
        self.indices['e100'] += 1
        parts = self._pad_line(raw_line, 'E100')
        # Adiciona √≠ndice para relacionamento com E110
        parts.append(self.indices['e100'])
        self.rows['E100'].append(parts)

    def _process_e110(self, raw_line: str) -> None:
        """
        Processa registro E110 (apura√ß√£o ICMS opera√ß√µes pr√≥prias).

        Atribui um √≠ndice sequencial e associa ao E100 mais recente.
        """
        self.indices['e110'] += 1
        parts = self._pad_line(raw_line, 'E110')
        # Adiciona √≠ndices: pr√≥prio e refer√™ncia ao E100
        parts.extend([self.indices['e110'], self.indices['e100']])
        self.rows['E110'].append(parts)

    def _process_e_child(self, registro: str, raw_line: str) -> None:
        """
        Processa registros filhos de E110 (E111, E112, E113, E115, E116).

        Adiciona √≠ndice do pai (E110).
        """
        parts = self._pad_line(raw_line, registro)
        parts.append(self.indices['e110'])
        self.rows[registro].append(parts)

    def _process_generic(self, registro: str, raw_line: str) -> None:
        """
        Processa um registro gen√©rico sem tratamento especial.

        Este m√©todo √© utilizado para registros de abertura (por exemplo, C001, D001, A001, F001, M001, E001) e
        quaisquer registros definidos em LAYOUTS que n√£o possuam l√≥gica espec√≠fica.
        Tamb√©m aplica l√≥gica de indexa√ß√£o din√¢mica se configurada.
        """
        parts = self._pad_line(raw_line, registro)
        
        # Aplica a√ß√µes de indexa√ß√£o din√¢mica
        if registro in self.record_actions:
            for action in self.record_actions[registro]:
                key = action['key']
                if action['type'] == 'increment':
                    # Garante que o contador existe
                    if key not in self.indices:
                        self.indices[key] = -1
                    self.indices[key] += 1
                    parts.append(str(self.indices[key]))
                elif action['type'] == 'read':
                    val = self.indices.get(key, -1)
                    parts.append(str(val))
        
        self.rows[registro].append(parts)

    # ======== Bloco C500 ========
    def _process_c500(self, raw_line: str) -> None:
        """Processa registro C500 (nota de energia/servi√ßo de comunica√ß√£o com cr√©dito)."""
        self.indices['c500'] += 1
        parts = self._pad_line(raw_line, 'C500')
        # Adiciona √≠ndices: pr√≥prio e refer√™ncia ao C010
        parts.extend([self.indices['c500'], self.indices['c010']])
        self.rows['C500'].append(parts)

    def _process_c5_child(self, registro: str, raw_line: str) -> None:
        """Processa registros filhos de C500 (C501, C505)."""
        parts = self._pad_line(raw_line, registro)
        # Adiciona √≠ndice do pai C500
        parts.append(self.indices['c500'])
        self.rows[registro].append(parts)

    # ======== Bloco D500 ========
    def _process_d500(self, raw_line: str) -> None:
        """Processa registro D500 (nota de servi√ßos de comunica√ß√£o com cr√©dito)."""
        self.indices['d500'] += 1
        parts = self._pad_line(raw_line, 'D500')
        # Adiciona √≠ndices: pr√≥prio e refer√™ncia ao D010
        parts.extend([self.indices['d500'], self.indices['d010']])
        self.rows['D500'].append(parts)

    def _process_d5_child(self, registro: str, raw_line: str) -> None:
        """Processa registros filhos de D500 (D501, D505)."""
        parts = self._pad_line(raw_line, registro)
        parts.append(self.indices['d500'])
        self.rows[registro].append(parts)

    # ======== Bloco D700 ========
    def _process_d700(self, raw_line: str) -> None:
        """Processa registro D700 (NFCom ‚Äì nota fiscal fatura eletr√¥nica)."""
        self.indices['d700'] += 1
        parts = self._pad_line(raw_line, 'D700')
        # Adiciona √≠ndices: pr√≥prio e refer√™ncia ao D010 (estabelecimento)
        parts.extend([self.indices['d700'], self.indices['d010']])
        self.rows['D700'].append(parts)
    
    def _pad_line(self, raw_line: str, registro: str) -> List[str]:
        """
        Faz o parse e padding de uma linha para o tamanho esperado.

        Args:
            raw_line: Linha bruta
            registro: C√≥digo do registro

        Returns:
            Lista de campos com padding
        """
        parts = parse_sped_line(raw_line)
        if registro not in self.layouts:
            return parts
            
        expected_len = len(self.layouts[registro])
        
        # Adiciona campos vazios se necess√°rio
        if len(parts) < expected_len:
            parts.extend([''] * (expected_len - len(parts)))
        
        # Trunca se houver campos extras
        return parts[:expected_len]
    
    def _create_dataframes(self) -> Dict[str, pd.DataFrame]:
        """
        Cria DataFrames a partir das linhas processadas.

        Returns:
            Dicion√°rio com DataFrames
        """
        dataframes = {}
        
        for code, data in self.rows.items():
            if not data:
                if code in self.layouts:
                    dataframes[code] = pd.DataFrame(columns=self.layouts[code])
                continue
            
            # Define colunas baseado no tipo de registro
            columns = self._get_columns_for_code(code)
            dataframes[code] = pd.DataFrame(data, columns=columns)
        
        return dataframes
    
    def _get_columns_for_code(self, code: str) -> List[str]:
        """
        Retorna as colunas apropriadas para um c√≥digo de registro.

        Args:
            code: C√≥digo do registro

        Returns:
            Lista de nomes de colunas
        """
        if code not in self.layouts:
            return []
            
        base_columns = self.layouts[code]
        
        # Usa configura√ß√£o din√¢mica de a√ß√µes para determinar colunas extras de √≠ndice
        if code in self.record_actions:
            extra_cols = [action['col'] for action in self.record_actions[code]]
            return base_columns + extra_cols
            
        return base_columns


# =========================
# PROCESSAMENTO DE DADOS
# =========================

class SpedDataProcessor:
    """Processa e transforma dados SPED."""
    
    @staticmethod
    def convert_dataframes(dataframes: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """
        Converte campos num√©ricos e de data nos DataFrames.

        Args:
            dataframes: Dicion√°rio com DataFrames

        Returns:
            Dicion√°rio com DataFrames convertidos
        """
        logger.info("Convertendo campos num√©ricos e datas")
        
        for code, df in dataframes.items():
            if df.empty:
                continue
            
            # Converte campos num√©ricos
            if code in NUMERIC_COLUMNS:
                convert_numeric_columns(df, NUMERIC_COLUMNS[code])
            
            # Converte e mapeia campos espec√≠ficos
            SpedDataProcessor._convert_specific_fields(code, df)
        
        return dataframes
    
    @staticmethod
    def _convert_specific_fields(code: str, df: pd.DataFrame) -> None:
        """
        Converte campos espec√≠ficos de cada registro.

        Args:
            code: C√≥digo do registro
            df: DataFrame a ser modificado
        """
        # Bloco C100
        if code == 'C100':
            if 'IND_OPER' in df.columns:
                df['IND_OPER'] = df['IND_OPER'].map(IND_OPER_MAP).fillna(df['IND_OPER'])
            if 'IND_EMIT' in df.columns:
                df['IND_EMIT'] = df['IND_EMIT'].map(IND_EMIT_MAP).fillna(df['IND_EMIT'])
            if 'IND_FRT' in df.columns:
                df['IND_FRT'] = df['IND_FRT'].map(IND_FRT_MAP).fillna(df['IND_FRT'])
            if 'DT_DOC' in df.columns:
                df['DT_DOC_DATE'] = pd.to_datetime(df['DT_DOC'], format='%Y%m%d', errors='coerce')
            if 'DT_E_S' in df.columns:
                df['DT_E_S_DATE'] = pd.to_datetime(df['DT_E_S'], format='%Y%m%d', errors='coerce')
        
        # Bloco D100
        elif code == 'D100':
            if 'IND_OPER' in df.columns:
                df['IND_OPER'] = df['IND_OPER'].map(IND_OPER_MAP).fillna(df['IND_OPER'])
            if 'IND_EMIT' in df.columns:
                df['IND_EMIT'] = df['IND_EMIT'].map(IND_EMIT_MAP).fillna(df['IND_EMIT'])
            if 'IND_FRT' in df.columns:
                df['IND_FRT'] = df['IND_FRT'].map(IND_FRT_MAP).fillna(df['IND_FRT'])
            if 'DT_DOC' in df.columns:
                df['DT_DOC_DATE'] = pd.to_datetime(df['DT_DOC'], format='%d%m%Y', errors='coerce')
            if 'DT_A_P' in df.columns:
                df['DT_A_P_DATE'] = pd.to_datetime(df['DT_A_P'], format='%d%m%Y', errors='coerce')
        
        # Bloco A100
        elif code == 'A100':
            if 'DT_DOC' in df.columns:
                df['DT_DOC_DATE'] = pd.to_datetime(df['DT_DOC'], format='%Y%m%d', errors='coerce')
            if 'DT_EXE_SERV' in df.columns:
                df['DT_EXE_SERV_DATE'] = pd.to_datetime(df['DT_EXE_SERV'], format='%Y%m%d', errors='coerce')

        # Bloco E100 - Per√≠odo de apura√ß√£o
        elif code == 'E100':
            if 'DT_INI' in df.columns:
                df['DT_INI_DATE'] = pd.to_datetime(df['DT_INI'], format='%Y%m%d', errors='coerce')
            if 'DT_FIN' in df.columns:
                df['DT_FIN_DATE'] = pd.to_datetime(df['DT_FIN'], format='%Y%m%d', errors='coerce')

        # Bloco E113 - Informa√ß√µes adicionais dos ajustes (doc fiscal)
        elif code == 'E113':
            if 'DT_DOC' in df.columns:
                df['DT_DOC_DATE'] = pd.to_datetime(df['DT_DOC'], format='%Y%m%d', errors='coerce')

        # Bloco E116 - Obriga√ß√µes a recolher
        elif code == 'E116':
            if 'DT_VCTO' in df.columns:
                df['DT_VCTO_DATE'] = pd.to_datetime(df['DT_VCTO'], format='%Y%m%d', errors='coerce')
            if 'MES_REF' in df.columns:
                # Tenta interpretar MES_REF (MMYYYY) como data no primeiro dia do m√™s
                df['MES_REF_DATE'] = pd.to_datetime(df['MES_REF'].str.zfill(6) + '01', format='%m%Y%d', errors='coerce')

        # Bloco C500
        elif code == 'C500':
            # Datas de emiss√£o e entrada
            if 'DT_DOC' in df.columns:
                df['DT_DOC_DATE'] = pd.to_datetime(df['DT_DOC'], format='%Y%m%d', errors='coerce')
            if 'DT_ENT' in df.columns:
                df['DT_ENT_DATE'] = pd.to_datetime(df['DT_ENT'], format='%Y%m%d', errors='coerce')

        # Bloco D500
        elif code == 'D500':
            # Mapeia indicadores
            if 'IND_OPER' in df.columns:
                df['IND_OPER'] = df['IND_OPER'].map(IND_OPER_MAP).fillna(df['IND_OPER'])
            if 'IND_EMIT' in df.columns:
                df['IND_EMIT'] = df['IND_EMIT'].map(IND_EMIT_MAP).fillna(df['IND_EMIT'])
            # Datas de emiss√£o e aquisi√ß√£o
            if 'DT_DOC' in df.columns:
                df['DT_DOC_DATE'] = pd.to_datetime(df['DT_DOC'], format='%Y%m%d', errors='coerce')
            if 'DT_A_P' in df.columns:
                df['DT_A_P_DATE'] = pd.to_datetime(df['DT_A_P'], format='%Y%m%d', errors='coerce')

        # Bloco D700
        elif code == 'D700':
            # Mapeia indicadores
            if 'IND_OPER' in df.columns:
                df['IND_OPER'] = df['IND_OPER'].map(IND_OPER_MAP).fillna(df['IND_OPER'])
            if 'IND_EMIT' in df.columns:
                df['IND_EMIT'] = df['IND_EMIT'].map(IND_EMIT_MAP).fillna(df['IND_EMIT'])
            # Datas de emiss√£o e entrada/sa√≠da
            if 'DT_DOC' in df.columns:
                df['DT_DOC_DATE'] = pd.to_datetime(df['DT_DOC'], format='%Y%m%d', errors='coerce')
            if 'DT_E_S' in df.columns:
                df['DT_E_S_DATE'] = pd.to_datetime(df['DT_E_S'], format='%Y%m%d', errors='coerce')
    
    @staticmethod
    def consolidate_group_new(
        dataframes: Dict[str, pd.DataFrame],
        parent_code: str,
        child_codes: List[str],
        parent_index_col: str,
        numeric_columns: Dict[str, List[str]] = None
    ) -> pd.DataFrame:
        """
        Consolida registros filhos com o registro pai (1-para-N).
        
        Diferente da vers√£o anterior que agrupava, esta vers√£o faz um LEFT JOIN,
        gerando uma linha para cada registro filho e repetindo os dados do pai.

        Args:
            dataframes: Dicion√°rio com DataFrames
            parent_code: C√≥digo do registro pai
            child_codes: Lista de c√≥digos dos registros filhos
            parent_index_col: Nome da coluna de √≠ndice do pai
            numeric_columns: (N√£o utilizado na nova l√≥gica de merge, mantido para compatibilidade)

        Returns:
            DataFrame consolidado com explos√£o de linhas
        """
        if parent_code not in dataframes or dataframes[parent_code].empty:
            return pd.DataFrame()

        # DEBUG VISUAL
        try:
            import streamlit as st
            st.toast(f"Consolidando {parent_code} com MERGE (V4.6 NEW)!", icon="üöÄ")
        except:
            pass
        
        # Prepara resultado inicial (apenas pai)
        result = dataframes[parent_code].copy().reset_index(drop=True)
        
        # Garante tipo str no √≠ndice do pai
        if parent_index_col in result.columns:
            result[parent_index_col] = result[parent_index_col].astype(str)
        
        for code in child_codes:
            child = dataframes.get(code)
            if child is None or child.empty:
                continue
            
            child = child.copy()
            
            # Garante tipo str no √≠ndice do filho
            if parent_index_col in child.columns:
                child[parent_index_col] = child[parent_index_col].astype(str)
            else:
                # Se filho n√£o tem a chave do pai, n√£o d√° pra juntar
                continue
            
            # Seleciona colunas para manter (exceto chaves e REG)
            keep_cols = [c for c in child.columns if c not in ['REG', parent_index_col]]
            
            if not keep_cols:
                continue

            # Renomeia colunas do filho com prefixo para evitar colis√£o
            rename_map = {c: f'{code}_{c}' for c in keep_cols}
            
            # Prepara dataframe do filho para o merge
            # Mant√©m parent_index_col para usar como chave
            child_to_merge = child[[parent_index_col] + keep_cols].rename(columns=rename_map)
            
            # Faz o Merge (Left Join)
            # Isso vai multiplicar as linhas do pai para cada linha do filho (Explos√£o)
            result = result.merge(child_to_merge, how='left', on=parent_index_col)
        
        return result
    
    @staticmethod
    def attach_header(
        df: pd.DataFrame,
        header_df: pd.DataFrame,
        parent_header_index_col: str,
        header_prefix: str
    ) -> pd.DataFrame:
        """
        Anexa informa√ß√µes de cabe√ßalho ao DataFrame.

        Args:
            df: DataFrame principal
            header_df: DataFrame com informa√ß√µes de cabe√ßalho
            parent_header_index_col: Coluna de √≠ndice do cabe√ßalho
            header_prefix: Prefixo para colunas do cabe√ßalho

        Returns:
            DataFrame com cabe√ßalho anexado
        """
        if header_df is None or header_df.empty or df.empty:
            return df
        
        # Seleciona colunas relevantes
        keep_cols = [c for c in header_df.columns if c not in ('REG', parent_header_index_col)]
        join_df = header_df[[parent_header_index_col] + keep_cols].rename(
            columns={c: f'{header_prefix}{c}' for c in keep_cols}
        )
        
        # Faz o merge e garante tipos compat√≠veis
        if parent_header_index_col in df.columns:
            df[parent_header_index_col] = df[parent_header_index_col].astype(str)
        if parent_header_index_col in join_df.columns:
            join_df[parent_header_index_col] = join_df[parent_header_index_col].astype(str)
            
        merged = df.merge(join_df, how='left', left_on=parent_header_index_col, right_on=parent_header_index_col)
        
        # Reordena colunas (cabe√ßalho primeiro)
        prefix_cols = [c for c in merged.columns if c.startswith(header_prefix)]
        other_cols = [c for c in merged.columns if c not in prefix_cols]
        
        return merged[prefix_cols + other_cols]


# =========================
# EXPORTA√á√ÉO PARA EXCEL
# =========================

class ExcelExporter:
    """Exporta dados para Excel com formata√ß√£o."""
    
    @staticmethod
    def should_format_as_currency(column_name: str) -> bool:
        """
        Verifica se uma coluna deve ser formatada como moeda.

        Args:
            column_name: Nome da coluna

        Returns:
            True se deve ser formatada como moeda
        """
        upper_name = str(column_name).upper()
        
        if upper_name.startswith('VL_') or 'VL_' in upper_name:
            return True
        
        currency_keywords = [
            'BC', 'ICMS', 'IPI', 'PIS', 'COFINS', 'SERV',
            'OPR', 'MERC', 'DESC', 'FRT', 'SEG', 'OUT', 'ISS'
        ]
        
        return any(keyword in upper_name for keyword in currency_keywords)
    
    @staticmethod
    def apply_currency_format(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
        """
        Aplica formata√ß√£o de moeda brasileira √†s colunas apropriadas.

        Args:
            writer: ExcelWriter do pandas
            df: DataFrame
            sheet_name: Nome da planilha
        """
        try:
            from openpyxl.utils import get_column_letter
            
            worksheet = writer.sheets[sheet_name]
            
            # Identifica colunas num√©ricas que devem ser formatadas
            numeric_cols = {c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])}
            currency_cols = {c for c in df.columns if ExcelExporter.should_format_as_currency(c)}
            cols_to_format = list(numeric_cols.union(currency_cols))
            
            for col_name in cols_to_format:
                if col_name not in df.columns:
                    continue
                
                col_index = df.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_index)
                
                # Aplica formato a todas as c√©lulas da coluna (exceto cabe√ßalho)
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f'{col_letter}{row}']
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = BRL_ACCOUNTING
        
        except Exception as e:
            logger.warning(f"Erro ao aplicar formata√ß√£o de moeda: {e}")
    
    @staticmethod
    def write_excel(sheets: Dict[str, pd.DataFrame], output_path: Path) -> None:
        """
        Escreve m√∫ltiplas planilhas em um arquivo Excel.

        Args:
            sheets: Dicion√°rio com nome da planilha e DataFrame
            output_path: Caminho do arquivo de sa√≠da
        """
        logger.info(f"Exportando para Excel: {output_path}")
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in sheets.items():
                    if df is None or df.empty:
                        logger.warning(f"Planilha {sheet_name} vazia, pulando")
                        continue
                    
                    # Escreve planilha
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Aplica formata√ß√£o de moeda em planilhas com valores
                    if (
                        sheet_name.endswith('CONSOLIDADO') or
                        sheet_name.startswith('M') or
                        sheet_name in [
                            # Bloco C detalhados
                            'C100', 'C170', 'C190', 'C197', 'C500', 'C501', 'C505',
                            # Bloco D detalhados
                            'D100', 'D170', 'D190', 'D101', 'D105',
                            'D500', 'D501', 'D505', 'D700',
                            # Bloco A e F
                            'A100', 'F100',
                            # Bloco E - Apura√ß√£o do ICMS/IPI
                            'E110', 'E111', 'E113', 'E115', 'E116'
                        ]
                    ):
                        ExcelExporter.apply_currency_format(writer, df, sheet_name)
            
            logger.info(f"Arquivo Excel criado com sucesso: {output_path}")
        
        except Exception as e:
            logger.error(f"Erro ao criar arquivo Excel: {e}")
            raise


# =========================
# PIPELINE PRINCIPAL
# =========================

class SpedProcessor:
    """Processador principal de arquivos SPED."""
    
    @staticmethod
    def process_single_file(input_path: Path) -> Dict[str, pd.DataFrame]:
        """
        Processa um √∫nico arquivo SPED.

        Args:
            input_path: Caminho do arquivo de entrada

        Returns:
            Dicion√°rio com planilhas consolidadas
        """
        logger.info(f"Processando arquivo: {input_path}")
        
        # Valida arquivo
        validate_file_path(input_path)
        
        # Parse
        parser = SpedParser(input_path)
        dataframes = parser.parse()
        
        # Converte campos
        dataframes = SpedDataProcessor.convert_dataframes(dataframes)
        
        # Consolida grupos
        consolidated = {}
        
        for group_name, (parent_code, child_codes, parent_idx, header_idx, header_code) in GROUPS.items():
            # Consolida filhos
            consolidated_df = SpedDataProcessor.consolidate_group_new(
                dataframes, parent_code, child_codes, parent_idx
            )
            
            if not consolidated_df.empty:
                # Anexa cabe√ßalho
                consolidated_df = SpedDataProcessor.attach_header(
                    consolidated_df,
                    dataframes.get(header_code),
                    header_idx,
                    f'{header_code}_'
                )
                
                # Remove colunas de √≠ndice
                consolidated_df.drop(
                    columns=[parent_idx, header_idx],
                    errors='ignore',
                    inplace=True
                )
            
            consolidated[f'{group_name}_CONSOLIDADO'] = consolidated_df
        
        return consolidated
    
    @staticmethod
    def process_multiple_files(input_paths: List[Path]) -> Dict[str, pd.DataFrame]:
        """
        Processa m√∫ltiplos arquivos SPED e consolida resultados.

        Args:
            input_paths: Lista de caminhos de arquivos

        Returns:
            Dicion√°rio com planilhas consolidadas
        """
        logger.info(f"Processando {len(input_paths)} arquivo(s)")
        
        results = []
        for path in input_paths:
            try:
                result = SpedProcessor.process_single_file(path)
                results.append(result)
            except Exception as e:
                logger.error(f"Erro ao processar {path}: {e}")
                raise
        
        # Concatena resultados
        return SpedProcessor._concat_results(results)
    
    @staticmethod
    def _concat_results(results: List[Dict[str, pd.DataFrame]]) -> Dict[str, pd.DataFrame]:
        """
        Concatena resultados de m√∫ltiplos arquivos.

        Args:
            results: Lista de dicion√°rios com DataFrames

        Returns:
            Dicion√°rio com DataFrames concatenados
        """
        all_keys = set().union(*[set(d.keys()) for d in results])
        concatenated = {}
        
        for key in all_keys:
            frames = [res[key] for res in results if key in res and not res[key].empty]
            concatenated[key] = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        
        return concatenated


# =========================
# INTERFACE DE LINHA DE COMANDO E GUI
# =========================

def main_cli(input_files: List[str], output_file: str) -> None:
    """
    Fun√ß√£o principal para interface de linha de comando.

    Args:
        input_files: Lista de caminhos de arquivos de entrada
        output_file: Caminho do arquivo de sa√≠da
    """
    try:
        input_paths = [Path(f) for f in input_files]
        output_path = Path(output_file)
        
        # Processa arquivos
        consolidated = SpedProcessor.process_multiple_files(input_paths)
        
        # Exporta para Excel
        ExcelExporter.write_excel(consolidated, output_path)
        
        logger.info("Processamento conclu√≠do com sucesso!")
    
    except Exception as e:
        logger.error(f"Erro no processamento: {e}")
        raise


def run_gui() -> None:
    """Executa interface gr√°fica com barra de progresso."""
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    import threading
    
    # Carrega configura√ß√µes da GUI
    window_title = get_config('gui.window_title', 'SPED ‚Üí Excel - Extrator de Contribui√ß√µes v3.0')
    window_size = get_config('gui.window_size', '800x450')
    default_output = get_config('gui.default_output_filename', 'sped_consolidado.xlsx')
    show_progress = get_config('gui.show_progress_bar', True)
    
    root = tk.Tk()
    root.title(window_title)
    root.geometry(window_size)
    
    input_paths = []
    output_var = tk.StringVar()
    status_var = tk.StringVar(value="Pronto para processar.")
    progress_var = tk.DoubleVar(value=0)
    
    def choose_input_files():
        """Abre di√°logo para selecionar arquivos de entrada."""
        nonlocal input_paths
        paths = filedialog.askopenfilenames(
            title="Selecione um ou mais arquivos SPED (.txt)",
            filetypes=[("Arquivos SPED", "*.txt"), ("Todos os arquivos", "*.*")]
        )
        if paths:
            input_paths = list(paths)
            label_selected.config(text=f"{len(input_paths)} arquivo(s) selecionado(s)")
            logger.info(f"Selecionados {len(input_paths)} arquivo(s)")
    
    def choose_output_file():
        """Abre di√°logo para selecionar arquivo de sa√≠da."""
        path = filedialog.asksaveasfilename(
            title="Salvar Excel como...",
            defaultextension=".xlsx",
            initialfile="sped_consolidado.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if path:
            output_var.set(path)
            logger.info(f"Arquivo de sa√≠da: {path}")
    
    def process_files():
        """Processa os arquivos selecionados."""
        if not input_paths:
            messagebox.showwarning("Aten√ß√£o", "Selecione pelo menos um arquivo SPED (.txt).")
            return
        
        if not output_var.get():
            messagebox.showwarning("Aten√ß√£o", "Escolha onde salvar o arquivo Excel (.xlsx).")
            return
        
        button_process.config(state="disabled")
        status_var.set("Processando arquivos...")
        root.update_idletasks()
        
        try:
            main_cli([str(p) for p in input_paths], output_var.get())
            status_var.set("Processamento conclu√≠do com sucesso!")
            messagebox.showinfo("Sucesso", f"Arquivo gerado:\n{output_var.get()}")
        
        except Exception as e:
            status_var.set("Erro no processamento.")
            messagebox.showerror("Erro", f"Erro ao processar arquivos:\n{str(e)}")
            logger.error(f"Erro na GUI: {e}")
        
        finally:
            button_process.config(state="normal")
    
    # Layout da interface
    frame = tk.Frame(root, padx=20, pady=20)
    frame.pack(fill="both", expand=True)
    
    # Se√ß√£o de arquivos de entrada
    tk.Label(frame, text="Arquivos SPED de entrada (.txt):", font=("Arial", 10, "bold")).grid(
        row=0, column=0, sticky="w", pady=(0, 5)
    )
    
    label_selected = tk.Label(frame, text="Nenhum arquivo selecionado", anchor="w")
    label_selected.grid(row=1, column=0, columnspan=2, sticky="we", pady=(0, 5))
    
    tk.Button(frame, text="Selecionar Arquivos .TXT", command=choose_input_files, width=25).grid(
        row=1, column=2, padx=10
    )
    
    # Se√ß√£o de arquivo de sa√≠da
    tk.Label(frame, text="Arquivo Excel de sa√≠da (.xlsx):", font=("Arial", 10, "bold")).grid(
        row=2, column=0, sticky="w", pady=(15, 5)
    )
    
    tk.Entry(frame, textvariable=output_var, width=70).grid(
        row=3, column=0, columnspan=2, sticky="we"
    )
    
    tk.Button(frame, text="Escolher Local", command=choose_output_file, width=25).grid(
        row=3, column=2, padx=10
    )
    
    # Bot√£o de processamento
    button_process = tk.Button(
        frame, text="Processar Arquivos", command=process_files,
        height=2, font=("Arial", 10, "bold"), bg="#4CAF50", fg="white"
    )
    button_process.grid(row=4, column=0, columnspan=3, sticky="we", pady=(20, 10))
    
    # Barra de progresso (se habilitada)
    if show_progress:
        progress_bar = ttk.Progressbar(
            frame, variable=progress_var, maximum=100, mode='determinate'
        )
        progress_bar.grid(row=5, column=0, columnspan=3, sticky="we", pady=(5, 10))
    
    # Barra de status
    tk.Label(frame, textvariable=status_var, anchor="w", relief="sunken", bd=1).grid(
        row=6, column=0, columnspan=3, sticky="we", pady=(5, 0)
    )
    
    frame.grid_columnconfigure(0, weight=1)
    
    root.mainloop()


# =========================
# PONTO DE ENTRADA
# =========================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Extrator de Contribui√ß√µes SPED v2.0 - Consolida blocos C, D, A, F e M",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos de uso:
  
  # Interface gr√°fica (padr√£o)
  python Extrat_contribuicoes_V2.py
  
  # Linha de comando - arquivo √∫nico
  python Extrat_contribuicoes_V2.py arquivo.txt --out saida.xlsx
  
  # Linha de comando - m√∫ltiplos arquivos
  python Extrat_contribuicoes_V2.py arquivo1.txt arquivo2.txt --out consolidado.xlsx
        """
    )
    
    parser.add_argument(
        "inputs",
        nargs="*",
        help="Um ou mais arquivos SPED (.txt) para processar"
    )
    
    parser.add_argument(
        "--out",
        required=False,
        help="Caminho do arquivo Excel de sa√≠da (.xlsx)"
    )
    
    parser.add_argument(
        "--log-level",
        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
        default='INFO',
        help="N√≠vel de log (padr√£o: INFO)"
    )
    
    args = parser.parse_args()
    
    # Configura n√≠vel de log
    logging.getLogger().setLevel(getattr(logging, args.log_level))
    
    # Executa CLI ou GUI
    if args.inputs and args.out:
        main_cli(args.inputs, args.out)
    else:
        run_gui()
